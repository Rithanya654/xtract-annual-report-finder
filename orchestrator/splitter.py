import csv
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from urllib.parse import urlparse
from openpyxl import load_workbook
from shared.models import CompanyRow

# Maps CompanyRow field → all known column header variants (lowercase)
COLUMN_ALIASES = {
    "agent":          ["agent name", "agent_name", "company", "company name", "name",
                       "bank name", "entity", "institution"],
    "country":        ["country", "countrycode", "country_code", "country code",
                       "nation", "region", "input_url_region"],
    "domain":         ["domain", "site", "company domain"],
    "nickname":       ["nickname", "nick", "short name", "repo_nickname", "short_name"],
    "period_type":    ["period_type", "period type", "report period",
                       "repo_reporttype", "reporttype", "report_type"],
    "statement_url":  ["statement_url", "statement url", "pdf_url", "pdf url"],
    "landing_url":    ["landing url", "landing page url", "landing_url",
                       "landingpageurl", "website", "url", "ir_url"],
    "auto_id":        ["autoid", "auto_id", "auto id", "id"],
    "client_id":      ["clientid", "client_id", "client id"],
    "source_url_id":  ["sourceurlid", "source_url_id", "source url id"],
    "expected_year":  ["expectedreportyear", "expected_report_year", "expected year",
                       "report_year", "reportyear", "year", "fiscal year",
                       "fiscal_year", "financial year", "financial_year"],
    "mob_id":         ["mobid", "mob_id", "mob id"],
    "is_valid":       ["isvalidforrun", "is_valid_for_run", "is_valid", "valid",
                       "validforrun", "active"],
}


def _normalize(header) -> str:
    return str(header or "").lower().strip().replace(" ", "_")


def _map_columns(headers: list) -> dict[str, int]:
    norm = [_normalize(h) for h in headers]
    col = {}
    for field, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(norm):
            # also try without underscores
            if h in aliases or h.replace("_", "") in [a.replace("_", "").replace(" ", "") for a in aliases]:
                col[field] = i
                break
    return col


def _cell(row, idx) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _is_valid_row(col_map: dict, row) -> bool:
    """Skip rows where isvalidforrun is explicitly N/False/0."""
    idx = col_map.get("is_valid")
    if idx is None:
        return True  # column absent → include all
    val = _cell(row, idx).lower()
    return val not in ("n", "no", "false", "0", "invalid", "skip")


def _extract_domain(url: str) -> str:
    if not url:
        return ""
    try:
        netloc = urlparse(url).netloc
        # strip www.
        return netloc.lstrip("www.") if netloc else ""
    except Exception:
        return ""


def _build_row(col_map: dict, row) -> CompanyRow | None:
    agent = _cell(row, col_map.get("agent"))
    if not agent:
        return None

    landing_url = _cell(row, col_map.get("landing_url"))
    domain = _cell(row, col_map.get("domain")) or _extract_domain(landing_url)

    return CompanyRow(
        agent=agent,
        country=_cell(row, col_map.get("country")),
        domain=domain,
        nickname=_cell(row, col_map.get("nickname")),
        period_type=_cell(row, col_map.get("period_type")) or "Year End",
        statement_url=_cell(row, col_map.get("statement_url")),
        landing_url=landing_url,
        auto_id=_cell(row, col_map.get("auto_id")),
        client_id=_cell(row, col_map.get("client_id")),
        source_url_id=_cell(row, col_map.get("source_url_id")),
        expected_year=_cell(row, col_map.get("expected_year")),
        mob_id=_cell(row, col_map.get("mob_id")),
        region=_cell(row, col_map.get("country")),  # reuse country/region
    )


def parse_spreadsheet(filepath: str) -> list[CompanyRow]:
    ext = filepath.lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls"):
        return _parse_xlsx(filepath)
    elif ext == "csv":
        return _parse_csv(filepath)
    raise ValueError(f"Unsupported file type: {ext}")


def _parse_xlsx(filepath: str) -> list[CompanyRow]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [h if h is not None else "" for h in next(rows_iter)]
    col_map = _map_columns(headers)

    results = []
    for row in rows_iter:
        if not _is_valid_row(col_map, row):
            continue
        r = _build_row(col_map, row)
        if r:
            results.append(r)
    return results


def _parse_csv(filepath: str) -> list[CompanyRow]:
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = [h if h else "" for h in next(reader)]
        col_map = _map_columns(headers)
        results = []
        for row in reader:
            if not _is_valid_row(col_map, row):
                continue
            r = _build_row(col_map, row)
            if r:
                results.append(r)
    return results


def split_into_chunks(rows: list, chunk_size: int = 200) -> list[list]:
    return [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]
